import sqlite3
import pandas as pd
from openpyxl import load_workbook

# Connexió a la base de dades SQLite
conn = sqlite3.connect('institut.db')

# Funció per generar l'horari del grup
def generar_horari_grup(nom_grup):
    # Obtindre les dades de l'horari del grup de la base de dades
    query = "SELECT * FROM Horari WHERE grup = ?"
    horari_grup = pd.read_sql_query(query, conn, params=(nom_grup,))

    # Carregar la plantilla de l'horari base
    horari_base = load_workbook('HorariBase.xlsx')
    full_name = f"{nom_grup}.xlsx"

    # Crear un nou fitxer d'Excel per a l'horari del grup
    horari_grup.to_excel(full_name, index=False, engine='openpyxl')

    # Copiar les dades del nou fitxer a la plantilla de l'horari base
    horari_grup_sheet = load_workbook(full_name)['Sheet1']
    horari_base_sheet = horari_base['Sheet1']

    for row in horari_grup_sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        horari_base_sheet.append(row)

    # Guardar l'horari del grup actualitzat
    horari_base.save(full_name)

    print(f"S'ha generat l'horari per al grup {nom_grup}.")

# Demanar el nom del grup a generar l'horari
nom_grup = input("Introdueix el nom del grup: ")

# Generar l'horari del grup
generar_horari_grup(nom_grup)

# Tancar la connexió a la base de dades
conn.close()
